"""Script para ler e analisar os arquivos Excel de controle de motoboy"""
from openpyxl import load_workbook
import sys

def analisar_excel_motoboy(caminho):
    """Analisa um arquivo Excel de controle de motoboy"""
    print(f"\n{'='*80}")
    print(f"ANALISANDO: {caminho}")
    print(f"{'='*80}\n")
    
    try:
        # Tenta ler o arquivo
        wb = load_workbook(caminho, data_only=True)
        
        print(f"📊 Abas encontradas: {wb.sheetnames}\n")
        
        # Analisa cada aba
        for sheet_name in wb.sheetnames:
            print(f"\n{'─'*80}")
            print(f"ABA: {sheet_name}")
            print(f"{'─'*80}")
            
            ws = wb[sheet_name]
            
            # Pega dimensões
            max_row = ws.max_row
            max_col = ws.max_column
            
            print(f"\n📏 Dimensões: {max_row} linhas x {max_col} colunas")
            
            # Pega cabeçalhos (primeira linha)
            print(f"\n� Colunas encontradas (primeira linha):")
            headers = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(1, col).value
                headers.append(cell_value)
                print(f"  {col}. {cell_value}")
            
            # Mostra primeiras 10 linhas
            print(f"\n🔍 Primeiras linhas (até 10):")
            linhas_para_mostrar = min(11, max_row)  # Cabeçalho + 10 linhas
            
            for row in range(1, linhas_para_mostrar + 1):
                linha_dados = []
                for col in range(1, max_col + 1):
                    valor = ws.cell(row, col).value
                    linha_dados.append(str(valor) if valor is not None else "")
                print(f"  Linha {row}: {' | '.join(linha_dados)}")
            
            # Conta linhas com dados
            linhas_com_dados = 0
            for row in range(2, max_row + 1):  # Pula cabeçalho
                tem_dado = False
                for col in range(1, max_col + 1):
                    if ws.cell(row, col).value is not None:
                        tem_dado = True
                        break
                if tem_dado:
                    linhas_com_dados += 1
            
            print(f"\n📊 Linhas com dados (excluindo cabeçalho): {linhas_com_dados}")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro ao ler arquivo: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    import os
    
    # Ajusta o caminho base
    base_path = os.path.dirname(os.path.abspath(__file__))
    
    arquivos = [
        os.path.join(base_path, "BACKUP_ORGANIZADO/01_EMPRESA_GRN_PIZZAS/OPERACIONAL/Entregas_Motoboy/MODELO MOTOBOY.xlsx"),
        os.path.join(base_path, "BACKUP_ORGANIZADO/01_EMPRESA_GRN_PIZZAS/OPERACIONAL/Entregas_Motoboy/MODELO MOTOBOY (1).xlsx")
    ]
    
    for arquivo in arquivos:
        if os.path.exists(arquivo):
            analisar_excel_motoboy(arquivo)
        else:
            print(f"❌ Arquivo não encontrado: {arquivo}")
        print("\n" + "="*80 + "\n")
