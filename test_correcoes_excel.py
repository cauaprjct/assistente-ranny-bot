"""
Teste das correções de Excel implementadas
Testa: estilos em linhas adicionadas e validação de limites
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'assistente-ranny'))

from openpyxl import load_workbook
from openpyxl.styles import Border
import io
import pdf_tools

passed = 0
failed = 0

def test(nome, condicao):
    global passed, failed
    if condicao:
        print(f"  ✅ {nome}")
        passed += 1
    else:
        print(f"  ❌ {nome}")
        failed += 1

print("=" * 60)
print("🧪 TESTE: Correções de Excel - assistente-ranny/pdf_tools.py")
print("=" * 60)

# ===== 1. Criar planilha base =====
print("\n📋 1. Criando planilha base...")
xlsx = pdf_tools.criar_xlsx_tabela(
    dados=[["João", 100, "01/01"], ["Maria", 200, "02/02"], ["Pedro", 300, "03/03"]],
    cabecalho=["Nome", "Valor", "Data"],
    titulo="Teste Correções"
)
test("Planilha criada", xlsx is not None)

if not xlsx:
    print("\n❌ Não foi possível criar planilha base. Abortando testes.")
    sys.exit(1)

# ===== 2. Adicionar linha COM estilo =====
print("\n✏️ 2. Testando adição de linha com estilo...")
xlsx2 = pdf_tools.editar_xlsx_adicionar_linha(xlsx, ["Ana", 400, "04/04"])
test("Linha adicionada", xlsx2 is not None)

if xlsx2:
    wb = load_workbook(io.BytesIO(xlsx2))
    ws = wb.active
    test("Total de linhas correto (5: header + 3 dados + 1 nova)", ws.max_row == 5)
    
    cell = ws.cell(row=5, column=1)
    test("Valor da célula correto", cell.value == "Ana")
    has_border = cell.border and cell.border.left and cell.border.left.style == 'thin'
    test("Célula tem borda thin", has_border)
    wb.close()

# ===== 3. Adicionar múltiplas linhas COM estilo =====
print("\n✏️ 3. Testando adição de múltiplas linhas com estilo...")
xlsx3 = pdf_tools.editar_xlsx_adicionar_linhas(xlsx, [["X", 1, "a"], ["Y", 2, "b"]])
test("Múltiplas linhas adicionadas", xlsx3 is not None)

if xlsx3:
    wb = load_workbook(io.BytesIO(xlsx3))
    ws = wb.active
    test("Total de linhas correto (6: header + 3 + 2 novas)", ws.max_row == 6)
    cell_x = ws.cell(row=5, column=1)
    cell_y = ws.cell(row=6, column=1)
    test("Primeira linha nova correta", cell_x.value == "X")
    test("Segunda linha nova correta", cell_y.value == "Y")
    has_border_x = cell_x.border and cell_x.border.left and cell_x.border.left.style == 'thin'
    test("Primeira linha nova tem borda", has_border_x)
    wb.close()

# ===== 4. Remover linha válida =====
print("\n🗑️ 4. Testando remoção de linha válida...")
xlsx4 = pdf_tools.editar_xlsx_remover_linha(xlsx, 3)  # Remove "Maria"
test("Linha removida", xlsx4 is not None)

if xlsx4:
    wb = load_workbook(io.BytesIO(xlsx4))
    ws = wb.active
    test("Total de linhas após remoção (3: header + João + Pedro)", ws.max_row == 3)
    test("Linha 2 agora é João", ws.cell(row=2, column=1).value == "João")
    test("Linha 3 agora é Pedro", ws.cell(row=3, column=1).value == "Pedro")
    wb.close()

# ===== 5. Proteção de limites =====
print("\n🛡️ 5. Testando proteção de limites...")
result_zero = pdf_tools.editar_xlsx_remover_linha(xlsx, 0)
test("Remover linha 0 retorna None", result_zero is None)

result_negative = pdf_tools.editar_xlsx_remover_linha(xlsx, -1)
test("Remover linha -1 retorna None", result_negative is None)

result_beyond = pdf_tools.editar_xlsx_remover_linha(xlsx, 999)
test("Remover linha 999 retorna None", result_beyond is None)

# ===== 6. Teste de integração: criar e editar =====
print("\n🔄 6. Teste de integração completo...")
# Cria planilha
xlsx_int = pdf_tools.criar_xlsx_tabela(
    dados=[["Item 1", 10], ["Item 2", 20]],
    cabecalho=["Descrição", "Quantidade"],
    titulo="Integração"
)
test("Planilha de integração criada", xlsx_int is not None)

if xlsx_int:
    # Adiciona linha
    xlsx_int = pdf_tools.editar_xlsx_adicionar_linha(xlsx_int, ["Item 3", 30])
    test("Linha adicionada na integração", xlsx_int is not None)
    
    if xlsx_int:
        # Substitui texto
        xlsx_int, subs = pdf_tools.editar_xlsx_substituir(xlsx_int, "Item", "Produto")
        test("Substituição realizada", xlsx_int is not None and subs == 3)
        
        if xlsx_int:
            # Remove linha
            xlsx_int = pdf_tools.editar_xlsx_remover_linha(xlsx_int, 2)
            test("Remoção na integração", xlsx_int is not None)
            
            if xlsx_int:
                # Verifica resultado final
                wb = load_workbook(io.BytesIO(xlsx_int))
                ws = wb.active
                test("Resultado final correto (3 linhas: header + 2 dados)", ws.max_row == 3)
                test("Substituição aplicada", "Produto" in str(ws.cell(row=2, column=1).value))
                wb.close()

# ===== Resultado =====
print("\n" + "=" * 60)
total = passed + failed
print(f"📊 Resultado: {passed}/{total} testes passaram")
if failed == 0:
    print("🎉 TODOS OS TESTES PASSARAM!")
    print("\n✅ As correções de Excel estão funcionando corretamente!")
    print("   - Estilos preservados ao adicionar linhas")
    print("   - Validação de limites ao remover linhas")
    print("   - Integração entre funções funcionando")
else:
    print(f"⚠️ {failed} teste(s) falharam")
print("=" * 60)

sys.exit(0 if failed == 0 else 1)
